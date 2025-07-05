import os
import base64
import cairosvg
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from tqdm import tqdm
from tempfile import NamedTemporaryFile
from test import (
    load_svg, encode_svg, parse_svg, compute_procrustes_similarity, quality_weighted_fusion, compare_svg_colors,
    collection, mongo_collection, TEMP_DIR
)
import numpy as np

QUERY_DIR = './dataset/Test_Dataset'  # Change as needed
OUTPUT_EXCEL = 'query_results.xlsx'
TOP_K = 7

os.makedirs(TEMP_DIR, exist_ok=True)

def svg_to_png_temp(svg_content):
    temp_file = NamedTemporaryFile(delete=False, suffix='.png')
    png_data = cairosvg.svg2png(bytestring=svg_content.encode('utf-8'))
    temp_file.write(png_data)
    temp_file.close()
    return temp_file.name

def find_similar_logos(query_svg_path, top_k=7):
    temp_path = query_svg_path  # No need to copy, just use path
    # DeepSVG vector
    deep_vector = None
    try:
        svg = load_svg(temp_path)
        deep_vector = encode_svg(svg).cpu().numpy().flatten()
    except Exception as e:
        print(f"[WARNING] DeepSVG encoding failed: {e}")

    # Algorithmic vector
    alg_vector = None
    try:
        alg_vector = parse_svg(temp_path)
    except Exception as e:
        print(f"[ERROR] Algorithmic parsing failed: {e}")
        return []

    deep_mongo_docs = []
    if deep_vector is not None:
        try:
            deep_results = collection.search(
                data=[deep_vector],
                anns_field="vector",
                param={"metric_type": "COSINE"},
                limit=10,
                output_fields=["milvus_id"]
            )[0]
            for hit in deep_results:
                doc = mongo_collection.find_one({"milvus_id": hit.id})
                if doc:
                    deep_mongo_docs.append({
                        "_id": str(doc["_id"]),
                        "score": float(hit.distance),
                        "doc": doc
                    })
        except Exception as e:
            print(f"[ERROR] Milvus search failed: {e}")

    # Algorithm search
    alg_matches = []
    for doc in mongo_collection.find({}, {"parsed_coordinates": 1, "svg_content": 1}):
        if "parsed_coordinates" not in doc or not doc["parsed_coordinates"]:
            continue
        score = compute_procrustes_similarity(alg_vector, np.array(doc["parsed_coordinates"]))
        alg_matches.append({
            "_id": str(doc["_id"]),
            "score": float(score),
            "doc": doc
        })
    alg_top_matches = sorted(alg_matches, key=lambda x: (-x["score"], x["_id"]))[:10]

    # Merge all unique IDs from both methods
    all_ids = set([doc["_id"] for doc in deep_mongo_docs] + [doc["_id"] for doc in alg_top_matches])
    fusion_candidates = []
    for doc_id in all_ids:
        deep_score = next((d["score"] for d in deep_mongo_docs if d["_id"] == doc_id), 0)
        alg_score = next((a["score"] for a in alg_matches if a["_id"] == doc_id), 0)
        doc = next((d["doc"] for d in deep_mongo_docs if d["_id"] == doc_id), None)
        if not doc:
            doc = next((a["doc"] for a in alg_matches if a["_id"] == doc_id), None)
        fusion_candidates.append({
            "_id": doc_id,
            "deep_score": deep_score,
            "alg_score": alg_score,
            "doc": doc
        })
    deep_scores = [c["deep_score"] for c in fusion_candidates]
    alg_scores = [c["alg_score"] for c in fusion_candidates]
    fused_2d_scores = quality_weighted_fusion(deep_scores, alg_scores)
    COLOR_WEIGHT = 0.2
    COLOR_THRESHOLD = 0.7
    for i, candidate in enumerate(fusion_candidates):
        doc_svg_content = candidate["doc"]["svg_content"]
        with NamedTemporaryFile(suffix=".svg", delete=False) as temp_svg_file:
            temp_svg_file.write(doc_svg_content.encode('utf-8'))
            temp_svg_file_path = temp_svg_file.name
        try:
            color_score = compare_svg_colors(temp_path, temp_svg_file_path)
        except Exception as e:
            print(f"[COLOR COMPARE ERROR] Failed for {candidate['_id']}: {e}")
            color_score = 0
        finally:
            if os.path.exists(temp_svg_file_path):
                os.remove(temp_svg_file_path)
        candidate["color_score"] = color_score
        if color_score >= COLOR_THRESHOLD:
            final_score = fused_2d_scores[i] + COLOR_WEIGHT * (color_score - COLOR_THRESHOLD)
        else:
            final_score = fused_2d_scores[i]
        candidate["fused_score"] = float(final_score)
    selected = sorted(fusion_candidates, key=lambda x: -x["fused_score"])[:top_k]
    return selected

def main():
    query_svgs = [f for f in os.listdir(QUERY_DIR) if f.lower().endswith('.svg')]
    results = []
    for svg_filename in tqdm(query_svgs, desc='Processing queries'):
        svg_path = os.path.join(QUERY_DIR, svg_filename)
        matches = find_similar_logos(svg_path, top_k=TOP_K)
        # Save query SVG as PNG for Excel
        with open(svg_path, 'r', encoding='utf-8') as f:
            query_svg_content = f.read()
        query_png_path = svg_to_png_temp(query_svg_content)
        # Save similar SVGs as PNGs for Excel
        similar_png_paths = []
        for item in matches:
            doc = item["doc"]
            try:
                png_path = svg_to_png_temp(doc["svg_content"])
                similar_png_paths.append(png_path)
            except Exception as e:
                print(f"[ERROR] Could not convert similar SVG to PNG: {e}")
                similar_png_paths.append(None)
        while len(similar_png_paths) < TOP_K:
            similar_png_paths.append(None)
        row = [svg_filename, query_png_path] + similar_png_paths[:TOP_K]
        results.append(row)
    columns = ["Filename", "Query Image"] + [f"Similar Image {i+1}" for i in range(TOP_K)]
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Similarity Results"
    ws.append(columns)
    image_size = 80
    cell_width = 15
    for row_idx, row in enumerate(results, start=2):
        ws.row_dimensions[row_idx].height = image_size * 0.75
        for col_idx, img_path in enumerate(row, start=1):
            if col_idx == 1:
                ws.cell(row=row_idx, column=col_idx, value=img_path)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = cell_width
            elif img_path and os.path.exists(img_path):
                img = ExcelImage(img_path)
                img.width, img.height = image_size, image_size
                ws.add_image(img, ws.cell(row=row_idx, column=col_idx).coordinate)
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = image_size // 6
    wb.save(OUTPUT_EXCEL)
    print(f"Results saved to {OUTPUT_EXCEL}")

if __name__ == '__main__':
    main() 