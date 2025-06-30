# =========================
# Standard Library Imports
# =========================
import os
import uuid
import io
import base64

# =========================
# Image Processing
# =========================
from PIL import Image as ImagePil
import cairosvg

# =========================
# Math, Data, and ML
# =========================
import numpy as np
from sklearn.decomposition import PCA
from scipy.spatial import procrustes
from scipy.spatial.distance import directed_hausdorff

# =========================
# DeepSVG
# =========================
from deepsvg.svglib.svg import SVG
from deepsvg import utils
from deepsvg.difflib.tensor import SVGTensor
from deepsvg.svglib.utils import to_gif
from deepsvg.svglib.geom import Bbox
from deepsvg.svgtensor_dataset import SVGTensorDataset, load_dataset
from deepsvg.utils.utils import batchify, linear
from configs.deepsvg.hierarchical_ordered import Config

# =========================
# Vector DB and DB Clients
# =========================
from pymilvus import connections, FieldSchema, DataType, CollectionSchema, Collection
import pymongo

# =========================
# SVG Utilities
# =========================
from svgpathtools import svg2paths, Path

# MongoDB Atlas URI (make sure to keep this secret!)
MONGO_URI = "mongodb+srv://hiru23anjalee:p24gomepFiz7R9HB@cluster0.kt9cubt.mongodb.net/?retryWrites=true&w=majority"

try:
    mongo_client = pymongo.MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
    mongo_client.server_info()  # Force connection check
    print(mongo_client.server_info())
    mongo_db = mongo_client["logoDBwc"]             # your DB name
    mongo_collection = mongo_db["logos"]          # your collection name
    print("✅ MongoDB Atlas connection established successfully.")
except pymongo.errors.ServerSelectionTimeoutError as err:
    print(f"❌ Failed to connect to MongoDB Atlas: {err}")

# Helper: Join all paths in SVG into one Path object
def order_points(points):
    center = np.mean(points, axis=0)
    angles = np.arctan2(points[:,1] - center[1], points[:,0] - center[0])
    return points[np.argsort(angles)]

def join_svg_paths(svg_file):
    paths, _ = svg2paths(svg_file)
    combined_path = Path()
    for path in paths:
        combined_path.extend(path)
    return combined_path


def parse_svg(svg_path, num_samples=250):
    path = join_svg_paths(svg_path)
    total_length = path.length()
    sample_distances = np.linspace(0, total_length, num_samples)
    sampled_points = []
    for distance in sample_distances:
        point = path.point(distance / total_length)
        sampled_points.append((point.real, point.imag))
    points = order_points(np.array(sampled_points))
    return points

# Helper: Compute Hausdorff similarity

def center_shape(points):
    centroid = np.mean(points, axis=0)
    return points - centroid

def scale_to_unit_size(points):
    min_coords = np.min(points, axis=0)
    max_coords = np.max(points, axis=0)
    width, height = max_coords - min_coords
    scale_factor = 1 / (max(width, height) if max(width, height) != 0 else 1)
    return points * scale_factor

def align_orientation(points):
    pca = PCA(n_components=2)
    pca.fit(points)
    principal_axis = pca.components_[0]
    angle = np.arctan2(principal_axis[1], principal_axis[0])
    rotation_matrix = np.array([
        [np.cos(-angle), -np.sin(-angle)],
        [np.sin(-angle), np.cos(-angle)]
    ])
    return points @ rotation_matrix.T

def normalize_shape(points):
    points = center_shape(points)
    points = align_orientation(points) 
    points = scale_to_unit_size(points)
    return points

def compute_hausdorff_similarity(shape1, shape2):
    shape1 = normalize_shape(shape1)
    shape2 = normalize_shape(shape2)
    forward_distance = directed_hausdorff(shape1, shape2)[0]
    reverse_distance = directed_hausdorff(shape2, shape1)[0]
    hausdorff_distance = max(forward_distance, reverse_distance)
    similarity = 1 / (1 + hausdorff_distance)
    return similarity

#############################################
# Batch Query and Excel Export Script
#############################################

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from tqdm import tqdm
from tempfile import NamedTemporaryFile

QUERY_DIR = './dataset/Test_Dataset'
OUTPUT_EXCEL = 'query_results.xlsx'

def svg_to_png_temp(svg_content):
    temp_file = NamedTemporaryFile(delete=False, suffix='.png')
    png_data = cairosvg.svg2png(bytestring=svg_content.encode('utf-8'))
    temp_file.write(png_data)
    temp_file.close()
    return temp_file.name

# Gather all SVG files in the query directory
query_svgs = [f for f in os.listdir(QUERY_DIR) if f.lower().endswith('.svg')]

results = []

for svg_filename in tqdm(query_svgs, desc='Processing queries'):
    svg_path = os.path.join(QUERY_DIR, svg_filename)
    try:
        alg_vector = parse_svg(svg_path)
    except Exception as e:
        print(f"[ERROR] Could not parse {svg_filename}: {e}")
        continue

    # Algorithm-based matching (Hausdorff)
    alg_matches = []
    for doc in mongo_collection.find({}, {"parsed_coordinates": 1, "svg_content": 1, "companyName": 1, "websiteURL": 1}):
        if "parsed_coordinates" not in doc or not doc["parsed_coordinates"]:
            continue
        score = compute_hausdorff_similarity(alg_vector, np.array(doc["parsed_coordinates"]))
        alg_matches.append({
            "_id": str(doc["_id"]),
            "score": float(score),
            "doc": doc
        })
    alg_top_matches = sorted(alg_matches, key=lambda x: (-x["score"], x["_id"]))[:5]

    # Save query SVG as PNG for Excel
    with open(svg_path, 'r', encoding='utf-8') as f:
        query_svg_content = f.read()
    query_png_path = svg_to_png_temp(query_svg_content)

    # Save similar SVGs as PNGs for Excel
    similar_png_paths = []
    for item in alg_top_matches:
        doc = item["doc"]
        try:
            png_path = svg_to_png_temp(doc["svg_content"])
            similar_png_paths.append(png_path)
        except Exception as e:
            print(f"[ERROR] Could not convert similar SVG to PNG: {e}")
            similar_png_paths.append(None)

    # Pad to 7 columns for similar images if needed
    while len(similar_png_paths) < 7:
        similar_png_paths.append(None)

    # Compose row: Filename, Query Image, Similar Image 1-7
    row = [svg_filename, query_png_path] + similar_png_paths[:7]
    results.append(row)

# Create DataFrame
columns = ["Filename", "Query Image"] + [f"Similar Image {i+1}" for i in range(7)]
df = pd.DataFrame(results, columns=columns)

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Image Similarity Results"

# Add column headers
ws.append(columns)

# Image size settings
image_size = 80  # Default height and width in pixels
cell_width = 15  # Approximate column width

# Add images to Excel
for row_idx, row in enumerate(results, start=2):  # Start from row 2
    ws.row_dimensions[row_idx].height = image_size * 0.75  # Adjust row height
    for col_idx, img_path in enumerate(row, start=1):  # Start from column A (1)
        if col_idx == 1:  # Filename column
            ws.cell(row=row_idx, column=col_idx, value=img_path)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = cell_width  # Set column width
        elif img_path and os.path.exists(img_path):  # Ensure image exists
            img = ExcelImage(img_path)
            img.width, img.height = image_size, image_size  # Resize images
            ws.add_image(img, ws.cell(row=row_idx, column=col_idx).coordinate)
            # Adjust column width based on image size
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = image_size // 6  # Adjust based on image width

# Save workbook
wb.save(OUTPUT_EXCEL)
print(f"Results saved to {OUTPUT_EXCEL}")